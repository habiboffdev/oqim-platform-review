from __future__ import annotations

import base64
import csv
import hashlib
import json
import mimetypes
from collections.abc import Awaitable, Callable
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any, Literal
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from pydantic import BaseModel, ConfigDict, Field
from pypdf import PdfReader

from app.modules.business_brain.contracts import BusinessBrainIndexRecordContract
from app.modules.business_brain.memory import BusinessBrainMemoryService
from app.modules.business_brain.memory_contracts import (
    MemoryFactWriteInput,
    SourceUnitContextualizationOutput,
)
from app.modules.business_brain.source_media_artifacts import (
    SourceMediaArtifactStore,
    SourceMediaArtifactWrite,
)
from app.modules.commercial_spine.contracts import LLMGatewayRequest
from app.modules.commercial_spine.llm_gateway import LLMGateway
from app.modules.commercial_spine.repository import CommercialSpineRepository
from app.modules.retrieval_core.indexing import RetrievalIndexEmbeddingService

SourceKind = Literal[
    "website",
    "pdf",
    "text",
    "telegram_channel",
    "screenshot",
    "voice_note",
    "spreadsheet",
    "past_conversation",
]
MediaKind = Literal["image", "video", "audio", "document"]


class OnboardingSourceIngestionModel(BaseModel):
    model_config = ConfigDict(extra="forbid")


class OnboardingSourceIngestionRequest(OnboardingSourceIngestionModel):
    schema_version: Literal["onboarding_source_ingestion_request.v1"] = (
        "onboarding_source_ingestion_request.v1"
    )
    workspace_id: int = Field(gt=0)
    source_ref: str = Field(min_length=1)
    source_kind: SourceKind
    source_payload: dict[str, Any] = Field(default_factory=dict)
    content_bytes: bytes | None = Field(default=None, exclude=True)
    content_base64: str | None = None
    correlation_id: str = Field(min_length=1)
    idempotency_key: str = Field(min_length=1)
    actor_ref: str = "onboarding_learner"
    embed_source_units: bool = False
    contextualize_source_units: bool = False


class SourceMediaAsset(OnboardingSourceIngestionModel):
    schema_version: Literal["source_media_asset.v1"] = "source_media_asset.v1"
    media_ref: str = Field(min_length=1)
    source_ref: str = Field(min_length=1)
    media_type: MediaKind
    origin: str = Field(min_length=1)
    url: str | None = None
    content_type: str | None = None
    byte_size: int | None = Field(default=None, ge=0)
    content_hash: str | None = None
    artifact_ref: str | None = None
    caption: str | None = None
    alt_text: str | None = None
    channel: str | None = None
    channel_id: str | None = None
    channel_message_id: str | None = None
    grouped_id: str | None = None
    page_number: int | None = Field(default=None, ge=1)
    metadata: dict[str, Any] = Field(default_factory=dict)


class OnboardingSourceIngestionResult(OnboardingSourceIngestionModel):
    schema_version: Literal["onboarding_source_ingestion_result.v1"] = (
        "onboarding_source_ingestion_result.v1"
    )
    source_fact_id: str
    source_units: list[BusinessBrainIndexRecordContract] = Field(default_factory=list)
    media_assets: list[SourceMediaAsset] = Field(default_factory=list)
    extracted_text_preview: str = ""
    degraded_reasons: list[str] = Field(default_factory=list)


@dataclass(frozen=True, slots=True)
class SourceFetchResult:
    content: bytes
    content_type: str
    final_url: str | None = None


FetchSource = Callable[[str], Awaitable[SourceFetchResult]]
TranscribeAudio = Callable[[bytes, str], Awaitable[str]]


class OnboardingSourceIngestionService:
    """Turns onboarding sources into Business Brain facts and source units."""

    def __init__(
        self,
        *,
        repository: CommercialSpineRepository,
        fetch_source: FetchSource | None = None,
        transcribe_audio: TranscribeAudio | None = None,
        media_artifact_store: SourceMediaArtifactStore | None = None,
        gateway: LLMGateway | None = None,
    ) -> None:
        self._repository = repository
        self._gateway = gateway or LLMGateway(repository=repository)
        self._memory = BusinessBrainMemoryService(repository=repository, gateway=self._gateway)
        self._fetch_source = fetch_source or _fetch_url
        self._transcribe_audio = transcribe_audio
        self._media_artifact_store = media_artifact_store

    async def ingest(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> OnboardingSourceIngestionResult:
        (
            extracted_text,
            metadata,
            degraded,
            media_assets,
            media_payloads,
        ) = await self._extract_source(request)
        media_assets = await self._store_media_artifacts(
            request=request,
            media_assets=media_assets,
            media_payloads=media_payloads,
        )
        chunks = _chunk_text(extracted_text)
        fact_id = f"business_source:{request.source_ref}:ingested"
        has_content = bool(chunks or media_assets)
        processing_state = "indexed" if has_content and not degraded else "degraded"
        source_refs = _unique([request.source_ref, *metadata.get("source_refs", [])])

        await self._memory.write_memory_fact(
            MemoryFactWriteInput(
                workspace_id=request.workspace_id,
                fact_id=fact_id,
                fact_type="business_source_fact",
                entity_ref=f"workspace:source:{request.source_ref}",
                value={
                    "kind": request.source_kind,
                    "input": dict(request.source_payload),
                    "metadata": metadata,
                    "processing": {
                        "state": processing_state,
                        "source_unit_count": len(chunks),
                        "source_media_count": len(media_assets),
                        "degraded_reasons": list(degraded),
                    },
                    "media_assets": [
                        asset.model_dump(mode="json") for asset in media_assets
                    ],
                    "text_preview": _preview(extracted_text),
                },
                source_refs=source_refs,
                source="onboarding",
                status="active" if has_content else "degraded",
                approval_state="confirmed" if has_content else "blocked",
                confidence=0.92 if has_content else 0.0,
                risk_tier="low" if has_content else "medium",
                correlation_id=request.correlation_id,
                idempotency_key=request.idempotency_key,
                actor_ref=request.actor_ref,
            )
        )

        source_units = await self._persist_source_units(
            request=request,
            fact_id=fact_id,
            chunks=chunks,
            source_refs=source_refs,
        )
        source_unit_degraded = _unique(
            [
                str(unit.degraded_reason)
                for unit in source_units
                if unit.degraded_reason
            ]
        )
        await self._persist_media_assets(
            request=request,
            source_fact_id=fact_id,
            media_assets=media_assets,
        )
        return OnboardingSourceIngestionResult(
            source_fact_id=fact_id,
            source_units=source_units,
            media_assets=media_assets,
            extracted_text_preview=_preview(extracted_text),
            degraded_reasons=_unique([*degraded, *source_unit_degraded]),
        )

    async def _extract_source(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> tuple[str, dict[str, Any], list[str], list[SourceMediaAsset], dict[str, bytes]]:
        if request.source_kind == "website":
            return await self._extract_website(request)
        if request.source_kind == "pdf":
            return self._extract_pdf(request)
        if request.source_kind == "telegram_channel":
            return self._extract_telegram_channel(request)
        if request.source_kind == "screenshot":
            return self._extract_screenshot(request)
        if request.source_kind == "voice_note":
            return await self._extract_voice_note(request)
        if request.source_kind == "spreadsheet":
            return self._extract_spreadsheet(request)
        if request.source_kind == "past_conversation":
            return self._extract_past_conversation(request)
        text = str(request.source_payload.get("text") or "")
        media_assets = _payload_media_assets(
            request.source_payload.get("media"),
            source_ref=request.source_ref,
        )
        return (
            _normalize_text(text),
            {"content_type": "text/plain", "source_refs": []},
            [] if text.strip() or media_assets else ["empty_source"],
            media_assets,
            {},
        )

    async def _extract_website(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> tuple[str, dict[str, Any], list[str], list[SourceMediaAsset], dict[str, bytes]]:
        url = str(request.source_payload.get("url") or "").strip()
        if not url:
            return "", {"source_refs": []}, ["missing_url"], [], {}
        try:
            fetched = await self._fetch_source(url)
        except Exception:
            return "", {"url": url, "source_refs": [url]}, ["fetch_failed"], [], {}
        raw_text = fetched.content.decode("utf-8", errors="ignore")
        json_extraction = _website_json_to_text_and_media(
            raw_text,
            content_type=fetched.content_type,
            base_url=fetched.final_url or url,
            source_ref=request.source_ref,
        )
        if json_extraction is not None:
            text, media_assets = json_extraction
            degraded = [] if text or media_assets else ["empty_source"]
            return (
                text,
                {
                    "url": url,
                    "final_url": fetched.final_url or url,
                    "content_type": fetched.content_type,
                    "source_refs": [url],
                    "structured_source": "shopify_products_json",
                },
                degraded,
                media_assets,
                {},
            )

        html = raw_text
        text = _html_to_text(html)
        media_assets = _html_media_assets(
            html,
            base_url=fetched.final_url or url,
            source_ref=request.source_ref,
        )
        degraded = [] if text or media_assets else ["empty_source"]
        return (
            text,
            {
                "url": url,
                "final_url": fetched.final_url or url,
                "content_type": fetched.content_type,
                "source_refs": [url],
            },
            degraded,
            media_assets,
            {},
        )

    def _extract_pdf(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> tuple[str, dict[str, Any], list[str], list[SourceMediaAsset], dict[str, bytes]]:
        raw = request.content_bytes
        if raw is None and request.content_base64:
            try:
                raw = base64.b64decode(request.content_base64)
            except ValueError:
                return "", {"source_refs": []}, ["invalid_base64"], [], {}
        if not raw:
            return "", {"source_refs": []}, ["missing_file_content"], [], {}
        try:
            reader = PdfReader(BytesIO(raw))
            text = "\n".join((page.extract_text() or "") for page in reader.pages)
        except Exception:
            return "", {"content_type": "application/pdf", "source_refs": []}, [
                "pdf_text_unavailable"
            ], [], {}
        normalized = _normalize_text(text)
        media_assets, media_degraded, media_payloads = _pdf_media_assets(
            reader,
            source_ref=request.source_ref,
        )
        degraded = [] if normalized or media_assets else ["empty_source"]
        degraded.extend(media_degraded)
        return (
            normalized,
            {
                "content_type": "application/pdf",
                "page_count": len(reader.pages),
                "source_refs": [],
            },
            degraded,
            media_assets,
            media_payloads,
        )

    def _extract_telegram_channel(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> tuple[str, dict[str, Any], list[str], list[SourceMediaAsset], dict[str, bytes]]:
        channel_id = str(
            request.source_payload.get("channel_id")
            or request.source_payload.get("channel")
            or request.source_ref
        )
        messages = request.source_payload.get("messages")
        if not isinstance(messages, list):
            return (
                "",
                {"source_refs": [request.source_ref]},
                ["missing_channel_messages"],
                [],
                {},
            )
        source_metadata = request.source_payload.get("metadata")
        metadata = dict(source_metadata) if isinstance(source_metadata, dict) else {}
        structured_source = str(metadata.get("structured_source") or "").strip()
        text_parts: list[str] = []
        media_assets: list[SourceMediaAsset] = []
        for index, raw_message in enumerate(messages):
            if not isinstance(raw_message, dict):
                continue
            message_id = str(raw_message.get("message_id") or raw_message.get("id") or index)
            text = str(raw_message.get("text") or raw_message.get("caption") or "").strip()
            if text:
                if structured_source == "shopify_products_json":
                    text_parts.append(text)
                else:
                    text_parts.append(f"{channel_id}/{message_id}: {text}")
            asset = _channel_message_media_asset(
                raw_message,
                source_ref=request.source_ref,
                channel_id=channel_id,
                index=index,
                caption=text or None,
            )
            if asset is not None:
                media_assets.append(asset)
        extracted = _normalize_text("\n".join(text_parts))
        degraded = [] if extracted or media_assets else ["empty_source"]
        return (
            extracted,
            {
                **metadata,
                "channel": "telegram_channel",
                "channel_id": channel_id,
                "message_count": len(messages),
                "source_refs": [request.source_ref, f"telegram_channel:{channel_id}"],
            },
            degraded,
            media_assets,
            {},
        )

    def _extract_screenshot(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> tuple[str, dict[str, Any], list[str], list[SourceMediaAsset], dict[str, bytes]]:
        raw = request.content_bytes
        if raw is None and request.content_base64:
            try:
                raw = base64.b64decode(request.content_base64)
            except ValueError:
                return "", {"source_refs": []}, ["invalid_base64"], [], {}
        caption = _optional_str(request.source_payload.get("caption"))
        content_type = (
            _optional_str(
                request.source_payload.get("content_type")
                or request.source_payload.get("mime_type")
            )
            or "image/png"
        )
        media_ref = str(
            request.source_payload.get("media_ref")
            or f"source_media:{request.source_ref}:screenshot:000"
        )
        media_payloads = {media_ref: raw} if raw else {}
        media_assets = [
            SourceMediaAsset(
                media_ref=media_ref,
                source_ref=request.source_ref,
                media_type="image",
                origin="uploaded_screenshot",
                content_type=content_type,
                byte_size=len(raw) if raw else None,
                content_hash=hashlib.sha256(raw).hexdigest() if raw else None,
                caption=caption,
                metadata={
                    "file_name": _optional_str(request.source_payload.get("file_name")),
                },
            )
        ]
        degraded = [] if raw else ["missing_file_content"]
        return (
            _normalize_text(caption or ""),
            {
                "content_type": content_type,
                "source_refs": [],
                "file_name": _optional_str(request.source_payload.get("file_name")),
            },
            degraded,
            media_assets,
            media_payloads,
        )

    async def _extract_voice_note(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> tuple[str, dict[str, Any], list[str], list[SourceMediaAsset], dict[str, bytes]]:
        raw = request.content_bytes
        if raw is None and request.content_base64:
            try:
                raw = base64.b64decode(request.content_base64)
            except ValueError:
                return "", {"source_refs": []}, ["invalid_base64"], [], {}
        content_type = (
            _optional_str(
                request.source_payload.get("content_type")
                or request.source_payload.get("mime_type")
            )
            or "audio/ogg"
        )
        transcript = _optional_str(
            request.source_payload.get("transcript")
            or request.source_payload.get("text")
        )
        degraded: list[str] = []
        if not transcript:
            if not raw:
                degraded.append("missing_file_content")
            else:
                try:
                    if self._transcribe_audio is not None:
                        transcript = await self._transcribe_audio(raw, content_type)
                    else:
                        transcript = await self._transcribe_voice_note_audio(
                            request=request,
                            audio_bytes=raw,
                            mime_type=content_type,
                        )
                except Exception:
                    degraded.append("voice_transcription_failed")
        normalized = _normalize_text(transcript or "")
        if not normalized and not degraded:
            degraded.append("empty_transcription")

        media_ref = str(
            request.source_payload.get("media_ref")
            or f"source_media:{request.source_ref}:voice:000"
        )
        media_payloads = {media_ref: raw} if raw else {}
        media_assets = [
            SourceMediaAsset(
                media_ref=media_ref,
                source_ref=request.source_ref,
                media_type="audio",
                origin="uploaded_voice_note",
                content_type=content_type,
                byte_size=len(raw) if raw else None,
                content_hash=hashlib.sha256(raw).hexdigest() if raw else None,
                caption=normalized or None,
                metadata={
                    "file_name": _optional_str(request.source_payload.get("file_name")),
                },
            )
        ]
        return (
            normalized,
            {
                "content_type": content_type,
                "source_refs": [],
                "file_name": _optional_str(request.source_payload.get("file_name")),
                "transcription_state": "ready" if normalized else "degraded",
            },
            degraded,
            media_assets,
            media_payloads,
        )

    async def _transcribe_voice_note_audio(
        self,
        *,
        request: OnboardingSourceIngestionRequest,
        audio_bytes: bytes,
        mime_type: str,
    ) -> str:
        from app.modules.extraction_runtime.media_semantics import normalize_voice_message

        normalized = await normalize_voice_message(
            audio_bytes,
            mime_type,
            gateway=self._gateway,
            workspace_id=request.workspace_id,
            correlation_id=f"source-intake-media:{request.workspace_id}:{request.source_ref}",
            source_refs=[request.source_ref],
        )
        return normalized.text

    def _extract_spreadsheet(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> tuple[str, dict[str, Any], list[str], list[SourceMediaAsset], dict[str, bytes]]:
        raw = request.content_bytes
        if raw is None and request.content_base64:
            try:
                raw = base64.b64decode(request.content_base64)
            except ValueError:
                return "", {"source_refs": []}, ["invalid_base64"], [], {}
        if not raw:
            return "", {"source_refs": []}, ["missing_file_content"], [], {}

        file_name = _optional_str(request.source_payload.get("file_name")) or ""
        content_type = (
            _optional_str(request.source_payload.get("content_type"))
            or _content_type_for_name(file_name)
            or "application/octet-stream"
        )
        try:
            headers, rows = _spreadsheet_rows(
                raw,
                content_type=content_type,
                file_name=file_name,
            )
        except Exception:
            return (
                "",
                {
                    "content_type": content_type,
                    "file_name": file_name,
                    "source_refs": [],
                },
                ["spreadsheet_parse_failed"],
                [],
                {},
            )

        normalized = _spreadsheet_rows_to_text(
            headers=headers,
            rows=rows,
            file_name=file_name,
        )
        degraded = [] if rows else ["empty_spreadsheet"]
        return (
            normalized,
            {
                "content_type": content_type,
                "file_name": file_name,
                "columns": headers,
                "row_count": len(rows),
                "source_refs": [],
                "normalized_source": "spreadsheet_rows",
            },
            degraded,
            [],
            {},
        )

    def _extract_past_conversation(
        self,
        request: OnboardingSourceIngestionRequest,
    ) -> tuple[str, dict[str, Any], list[str], list[SourceMediaAsset], dict[str, bytes]]:
        turns = request.source_payload.get("turns")
        if not isinstance(turns, list):
            return "", {"source_refs": [request.source_ref]}, ["missing_conversation_turns"], [], {}
        text_parts: list[str] = []
        source_refs = [request.source_ref]
        for index, raw_turn in enumerate(turns):
            if not isinstance(raw_turn, dict):
                continue
            sender = str(raw_turn.get("sender_type") or raw_turn.get("role") or "unknown")
            content = str(raw_turn.get("content") or raw_turn.get("text") or "").strip()
            if not content:
                continue
            message_ref = str(
                raw_turn.get("message_ref")
                or raw_turn.get("id")
                or f"{request.source_ref}:turn:{index:03d}"
            )
            source_refs.append(message_ref)
            created_at = str(raw_turn.get("created_at") or "").strip()
            quality = str(raw_turn.get("quality_label") or "").strip()
            line_parts = [f"{message_ref}", f"sender={sender}"]
            if created_at:
                line_parts.append(f"created_at={created_at}")
            if quality:
                line_parts.append(f"quality={quality}")
            text_parts.append(f"{' '.join(line_parts)}: {content}")
        extracted = _normalize_text("\n".join(text_parts))
        return (
            extracted,
            {
                "content_type": "application/vnd.oqim.past-conversation+json",
                "turn_count": len(turns),
                "normalized_source": "past_conversation_turns",
                "conversation_id": request.source_payload.get("conversation_id"),
                "source_refs": _unique(source_refs),
            },
            [] if extracted else ["empty_source"],
            [],
            {},
        )

    async def _store_media_artifacts(
        self,
        *,
        request: OnboardingSourceIngestionRequest,
        media_assets: list[SourceMediaAsset],
        media_payloads: dict[str, bytes],
    ) -> list[SourceMediaAsset]:
        if self._media_artifact_store is None or not media_payloads:
            return media_assets
        stored_assets: list[SourceMediaAsset] = []
        for asset in media_assets:
            payload = media_payloads.get(asset.media_ref)
            if not payload:
                stored_assets.append(asset)
                continue
            stored = await self._media_artifact_store.write(
                SourceMediaArtifactWrite(
                    workspace_id=request.workspace_id,
                    media_ref=asset.media_ref,
                    content_bytes=payload,
                    content_type=asset.content_type,
                )
            )
            if stored is None:
                stored_assets.append(asset)
                continue
            stored_assets.append(
                asset.model_copy(
                    update={
                        "artifact_ref": stored.artifact_ref,
                        "byte_size": stored.byte_size,
                        "content_hash": stored.content_hash,
                        "content_type": stored.content_type or asset.content_type,
                    }
                )
            )
        return stored_assets

    async def _persist_source_units(
        self,
        *,
        request: OnboardingSourceIngestionRequest,
        fact_id: str,
        chunks: list[str],
        source_refs: list[str],
    ) -> list[BusinessBrainIndexRecordContract]:
        records: list[BusinessBrainIndexRecordContract] = []
        embedding_indexer = RetrievalIndexEmbeddingService()
        for index, chunk in enumerate(chunks):
            source_text = _contextual_source_text(
                request=request,
                index=index,
                chunk=chunk,
            )
            contextualization_reason: str | None = None
            if request.contextualize_source_units:
                source_text, contextualization_reason = await self._contextualized_source_unit_text(
                    request=request,
                    index=index,
                    fact_id=fact_id,
                    source_refs=source_refs,
                    source_text=source_text,
                )
            embedding_result = await embedding_indexer.embed_text(
                source_text,
                enabled=request.embed_source_units,
                context=f"onboarding_source:{request.source_ref}:{index}",
            )
            reason = embedding_result.degraded_reason or contextualization_reason
            state = "degraded" if embedding_result.degraded_reason else "ready"
            record = BusinessBrainIndexRecordContract(
                index_id=f"source:{fact_id}:unit:{index:03d}",
                workspace_id=request.workspace_id,
                fact_id=fact_id,
                unit_ref=f"source_unit:{fact_id}:{index:03d}",
                state=state,
                embedding_ref=(
                    f"{embedding_result.embedding_model}:{fact_id}:{index:03d}"
                    if embedding_result.embedding_model
                    else None
                ),
                embedding_model=embedding_result.embedding_model,
                embedding_state=embedding_result.embedding_state,
                embedding=embedding_result.embedding,
                source_text=source_text,
                degraded_reason=reason,
                source_refs=source_refs,
                idempotency_key=f"source:{request.idempotency_key}:unit:{index:03d}",
            )
            await self._repository.persist_index_record(record)
            records.append(record)
        return records

    async def _contextualized_source_unit_text(
        self,
        *,
        request: OnboardingSourceIngestionRequest,
        index: int,
        fact_id: str,
        source_refs: list[str],
        source_text: str,
    ) -> tuple[str, str | None]:
        prompt_id = "business_brain.source_unit_contextualization"
        prompt_version = "1.0.0"
        result = await self._gateway.generate(
            LLMGatewayRequest(
                route_key="structured_fast",
                workflow_name="onboarding.source_unit_contextualization",
                prompt_id=prompt_id,
                prompt_version=prompt_version,
                input_payload={
                    "source_kind": request.source_kind,
                    "source_ref": request.source_ref,
                    "source_payload": dict(request.source_payload),
                    "chunk_index": index,
                    "source_text": source_text[:6000],
                },
                output_schema_name="SourceUnitContextualizationOutput",
                workspace_id=request.workspace_id,
                correlation_id=f"source-intake-context:{request.workspace_id}:{fact_id}:{index}",
                source_refs=_unique([*source_refs, f"fact:{fact_id}"]),
                budget={"max_output_chars": 1200},
                timeout_ms=10_000,
                fallback_policy=["use_deterministic_context"],
            ),
            output_model=SourceUnitContextualizationOutput,
        )
        if result.status != "ok" or result.parsed_output is None:
            return source_text, f"contextualization:{result.status}"
        context = " ".join(str(result.parsed_output.get("context") or "").split())
        if not context:
            return source_text, "contextualization:empty"
        return _source_text_with_llm_context(context=context, source_text=source_text), None

    async def _persist_media_assets(
        self,
        *,
        request: OnboardingSourceIngestionRequest,
        source_fact_id: str,
        media_assets: list[SourceMediaAsset],
    ) -> None:
        for index, asset in enumerate(media_assets):
            source_refs = _unique([request.source_ref, asset.media_ref])
            await self._memory.write_memory_fact(
                MemoryFactWriteInput(
                    workspace_id=request.workspace_id,
                    fact_id=f"business_source_media:{asset.media_ref}",
                    fact_type="business_source_media_fact",
                    entity_ref=f"workspace:source_media:{asset.media_ref}",
                    value={
                        **asset.model_dump(mode="json"),
                        "source_fact_id": source_fact_id,
                    },
                    source_refs=source_refs,
                    source="onboarding",
                    status="active",
                    approval_state="confirmed",
                    confidence=0.9,
                    risk_tier="low",
                    correlation_id=request.correlation_id,
                    idempotency_key=f"source:{request.idempotency_key}:media:{index:03d}",
                    actor_ref=request.actor_ref,
                )
            )


async def _fetch_url(url: str) -> SourceFetchResult:
    async with httpx.AsyncClient(follow_redirects=True, timeout=20.0) as client:
        response = await client.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (compatible; OQIMBot/1.0)",
                "Accept": (
                    "text/html,application/xhtml+xml,application/json,"
                    "application/pdf,text/plain"
                ),
            },
        )
        response.raise_for_status()
        return SourceFetchResult(
            content=response.content,
            content_type=response.headers.get("content-type", "application/octet-stream"),
            final_url=str(response.url),
        )


def _html_to_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "nav", "footer", "header", "noscript"]):
        tag.decompose()
    title = soup.title.get_text(" ", strip=True) if soup.title else ""
    body = soup.get_text("\n", strip=True)
    return _normalize_text("\n".join(part for part in (title, body) if part))


def _html_media_assets(
    html: str,
    *,
    base_url: str,
    source_ref: str,
) -> list[SourceMediaAsset]:
    soup = BeautifulSoup(html, "html.parser")
    assets: list[SourceMediaAsset] = []
    seen: set[str] = set()

    def add_asset(
        *,
        raw_url: str | None,
        origin: str,
        media_type: MediaKind,
        alt_text: str | None = None,
        caption: str | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> None:
        if not raw_url:
            return
        url = urljoin(base_url, raw_url.strip())
        if not url or url in seen:
            return
        seen.add(url)
        content_type = _content_type_for_name(url)
        assets.append(
            SourceMediaAsset(
                media_ref=f"source_media:{source_ref}:{len(assets):03d}",
                source_ref=source_ref,
                media_type=media_type,
                origin=origin,
                url=url,
                content_type=content_type,
                alt_text=alt_text or None,
                caption=caption or None,
                metadata=metadata or {},
            )
        )

    for meta in soup.find_all("meta"):
        prop = str(meta.get("property") or meta.get("name") or "").strip().lower()
        if prop in {"og:image", "twitter:image"}:
            add_asset(
                raw_url=str(meta.get("content") or ""),
                origin=prop,
                media_type="image",
            )

    for image in soup.find_all("img"):
        add_asset(
            raw_url=str(image.get("src") or image.get("data-src") or ""),
            origin="html_img",
            media_type="image",
            alt_text=str(image.get("alt") or "").strip() or None,
            metadata={
                key: str(image.get(key))
                for key in ("width", "height")
                if image.get(key) is not None
            },
        )

    for video in soup.find_all("video"):
        add_asset(
            raw_url=str(video.get("src") or ""),
            origin="html_video",
            media_type="video",
        )
        for source in video.find_all("source"):
            add_asset(
                raw_url=str(source.get("src") or ""),
                origin="html_video_source",
                media_type="video",
                metadata={"type": str(source.get("type") or "")},
            )

    for audio in soup.find_all("audio"):
        add_asset(
            raw_url=str(audio.get("src") or ""),
            origin="html_audio",
            media_type="audio",
        )
        for source in audio.find_all("source"):
            add_asset(
                raw_url=str(source.get("src") or ""),
                origin="html_audio_source",
                media_type="audio",
                metadata={"type": str(source.get("type") or "")},
            )

    return assets[:100]


def _website_json_to_text_and_media(
    raw_text: str,
    *,
    content_type: str,
    base_url: str,
    source_ref: str,
) -> tuple[str, list[SourceMediaAsset]] | None:
    content_type_lower = content_type.lower()
    stripped = raw_text.strip()
    if "json" not in content_type_lower and not stripped.startswith(("{", "[")):
        return None
    try:
        payload = json.loads(stripped)
    except json.JSONDecodeError:
        return None
    products = _shopify_products(payload)
    if not products:
        return None

    text_parts: list[str] = []
    assets: list[SourceMediaAsset] = []
    seen_urls: set[str] = set()
    for product_index, product in enumerate(products[:50]):
        title = _optional_str(product.get("title"))
        handle = _optional_str(product.get("handle"))
        vendor = _optional_str(product.get("vendor"))
        product_type = _optional_str(product.get("product_type"))
        body = _html_to_text(str(product.get("body_html") or ""))
        options = _shopify_options_text(product.get("options"))
        variants = _shopify_variants_text(product.get("variants"))
        text_parts.append(
            _normalize_text(
                "\n".join(
                    part
                    for part in (
                        f"Product: {title}" if title else "",
                        f"Handle: {handle}" if handle else "",
                        f"Vendor: {vendor}" if vendor else "",
                        f"Type: {product_type}" if product_type else "",
                        body,
                        options,
                        variants,
                    )
                    if part
                )
            )
        )
        for image_index, image in enumerate(_shopify_images(product.get("images"))):
            url = _optional_str(image.get("src") or image.get("url"))
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
            image_id = _optional_str(image.get("id")) or f"{product_index}-{image_index}"
            assets.append(
                SourceMediaAsset(
                    media_ref=(
                        f"source_media:{source_ref}:shopify:"
                        f"{_slug_for_ref(str(product.get('id') or product_index))}:"
                        f"{_slug_for_ref(image_id)}"
                    ),
                    source_ref=source_ref,
                    media_type="image",
                    origin="shopify_product_image",
                    url=urljoin(base_url, url),
                    content_type=_content_type_for_name(url),
                    alt_text=title,
                    metadata={
                        "product_id": str(product.get("id") or ""),
                        "product_handle": handle or "",
                        "image_id": image_id,
                        "position": str(image.get("position") or image_index + 1),
                        "width": str(image.get("width") or ""),
                        "height": str(image.get("height") or ""),
                    },
                )
            )
    return _normalize_text("\n\n".join(part for part in text_parts if part)), assets[:100]


def _shopify_products(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, dict):
        raw_products = payload.get("products")
        if isinstance(raw_products, list):
            return [item for item in raw_products if isinstance(item, dict)]
        raw_product = payload.get("product")
        if isinstance(raw_product, dict):
            return [raw_product]
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict) and "title" in item]
    return []


def _shopify_images(raw_images: Any) -> list[dict[str, Any]]:
    if not isinstance(raw_images, list):
        return []
    return [image for image in raw_images if isinstance(image, dict)]


def _shopify_options_text(raw_options: Any) -> str:
    if not isinstance(raw_options, list):
        return ""
    lines: list[str] = []
    for option in raw_options:
        if not isinstance(option, dict):
            continue
        name = _optional_str(option.get("name"))
        values = option.get("values")
        value_text = ", ".join(str(value) for value in values) if isinstance(values, list) else ""
        if name or value_text:
            lines.append(f"Option: {name or 'variant'} = {value_text}".strip())
    return "\n".join(lines)


def _shopify_variants_text(raw_variants: Any) -> str:
    if not isinstance(raw_variants, list):
        return ""
    lines: list[str] = []
    for variant in raw_variants[:100]:
        if not isinstance(variant, dict):
            continue
        title = _optional_str(variant.get("title")) or "Default"
        sku = _optional_str(variant.get("sku"))
        price = _optional_str(variant.get("price"))
        available = variant.get("available")
        availability = (
            "available" if available is True else "unavailable" if available is False else ""
        )
        option_values = [
            _optional_str(variant.get(key))
            for key in ("option1", "option2", "option3")
            if _optional_str(variant.get(key))
        ]
        lines.append(
            "Variant: "
            + "; ".join(
                part
                for part in (
                    title,
                    f"sku={sku}" if sku else "",
                    f"price={price}" if price else "",
                    f"availability={availability}" if availability else "",
                    f"options={', '.join(option_values)}" if option_values else "",
                )
                if part
            )
        )
    return "\n".join(lines)


def _pdf_media_assets(
    reader: PdfReader,
    *,
    source_ref: str,
) -> tuple[list[SourceMediaAsset], list[str], dict[str, bytes]]:
    assets: list[SourceMediaAsset] = []
    degraded: list[str] = []
    media_payloads: dict[str, bytes] = {}
    try:
        pages = list(reader.pages)
    except Exception:
        return assets, ["pdf_media_unavailable"], media_payloads
    for page_index, page in enumerate(pages):
        try:
            page_images = list(page.images)
        except Exception:
            degraded.append("pdf_page_media_unavailable")
            continue
        for image_index, image in enumerate(page_images):
            name = str(getattr(image, "name", "") or f"image-{image_index}")
            data = getattr(image, "data", None)
            image_bytes = bytes(data or b"")
            media_ref = f"source_media:{source_ref}:pdf:{page_index + 1}:{image_index:03d}"
            if image_bytes:
                media_payloads[media_ref] = image_bytes
            assets.append(
                SourceMediaAsset(
                    media_ref=media_ref,
                    source_ref=source_ref,
                    media_type="image",
                    origin="pdf_page_image",
                    content_type=_content_type_for_name(name),
                    byte_size=len(image_bytes) if image_bytes else None,
                    content_hash=(
                        hashlib.sha256(image_bytes).hexdigest() if image_bytes else None
                    ),
                    page_number=page_index + 1,
                    metadata={"name": name},
                )
            )
    assets = assets[:100]
    return assets, _unique(degraded), {
        asset.media_ref: media_payloads[asset.media_ref]
        for asset in assets
        if asset.media_ref in media_payloads
    }


def _payload_media_assets(
    media_items: Any,
    *,
    source_ref: str,
) -> list[SourceMediaAsset]:
    if not isinstance(media_items, list):
        return []
    assets: list[SourceMediaAsset] = []
    for index, item in enumerate(media_items):
        if not isinstance(item, dict):
            continue
        media_type = _media_kind(item.get("media_type") or item.get("type"))
        if media_type is None:
            continue
        url = _optional_str(item.get("url") or item.get("media_url"))
        assets.append(
            SourceMediaAsset(
                media_ref=str(item.get("media_ref") or f"source_media:{source_ref}:{index:03d}"),
                source_ref=source_ref,
                media_type=media_type,
                origin=str(item.get("origin") or "payload_media"),
                url=url,
                content_type=_optional_str(item.get("content_type") or item.get("mime_type")),
                caption=_optional_str(item.get("caption")),
                alt_text=_optional_str(item.get("alt_text")),
                metadata={
                    key: value
                    for key, value in item.items()
                    if key
                    not in {
                        "media_ref",
                        "media_type",
                        "type",
                        "url",
                        "media_url",
                        "content_type",
                        "mime_type",
                        "caption",
                        "alt_text",
                        "origin",
                    }
                },
            )
        )
    return assets


def _channel_message_media_asset(
    message: dict[str, Any],
    *,
    source_ref: str,
    channel_id: str,
    index: int,
    caption: str | None,
) -> SourceMediaAsset | None:
    media_type = _media_kind(message.get("media_type") or message.get("mediaType"))
    if media_type is None:
        return None
    metadata = (
        dict(message.get("media_metadata"))
        if isinstance(message.get("media_metadata"), dict)
        else (
            dict(message.get("mediaMetadata"))
            if isinstance(message.get("mediaMetadata"), dict)
            else {}
        )
    )
    message_id = str(message.get("message_id") or message.get("id") or index)
    url = _optional_str(
        message.get("url")
        or message.get("media_url")
        or metadata.get("url")
        or metadata.get("media_url")
        or metadata.get("thumbnail_url")
    )
    content_type = _optional_str(
        message.get("content_type")
        or message.get("mime_type")
        or metadata.get("mime_type")
        or metadata.get("mimeType")
    )
    return SourceMediaAsset(
        media_ref=str(
            message.get("media_ref")
            or f"telegram_channel:{channel_id}:{message_id}:media"
        ),
        source_ref=source_ref,
        media_type=media_type,
        origin="telegram_channel_message",
        url=url,
        content_type=content_type,
        caption=caption,
        channel="telegram_channel",
        channel_id=channel_id,
        channel_message_id=message_id,
        grouped_id=_optional_str(message.get("grouped_id") or metadata.get("grouped_id")),
        metadata=metadata,
    )


def _chunk_text(text: str, *, max_chars: int = 1800) -> list[str]:
    paragraphs = [part.strip() for part in text.splitlines() if part.strip()]
    chunks: list[str] = []
    current = ""
    for paragraph in paragraphs:
        candidate = f"{current}\n{paragraph}".strip() if current else paragraph
        if len(candidate) <= max_chars:
            current = candidate
            continue
        if current:
            chunks.append(current)
        current = paragraph[:max_chars]
    if current:
        chunks.append(current)
    return chunks[:50]


def _normalize_text(text: str) -> str:
    lines = [" ".join(line.split()) for line in text.splitlines()]
    return "\n".join(line for line in lines if line)


def _spreadsheet_rows(
    raw: bytes,
    *,
    content_type: str,
    file_name: str,
) -> tuple[list[str], list[list[str]]]:
    normalized_content_type = content_type.lower()
    normalized_name = file_name.lower()
    if (
        normalized_content_type in {"text/csv", "application/csv"}
        or normalized_name.endswith(".csv")
    ):
        return _csv_rows(raw)
    return _xlsx_rows(raw)


def _csv_rows(raw: bytes) -> tuple[list[str], list[list[str]]]:
    text = raw.decode("utf-8-sig")
    rows = list(csv.reader(StringIO(text)))
    return _normalize_spreadsheet_rows(rows)


def _xlsx_rows(raw: bytes) -> tuple[list[str], list[list[str]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=BytesIO(raw),
        read_only=True,
        data_only=True,
    )
    try:
        worksheet = workbook.active
        if worksheet is None:
            return [], []
        rows = [
            [_spreadsheet_cell_value(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        return _normalize_spreadsheet_rows(rows)
    finally:
        workbook.close()


def _normalize_spreadsheet_rows(rows: list[list[Any]]) -> tuple[list[str], list[list[str]]]:
    non_empty = [
        [_spreadsheet_cell_value(cell) for cell in row]
        for row in rows
        if any(_spreadsheet_cell_value(cell) for cell in row)
    ]
    if not non_empty:
        return [], []
    headers = [
        value or f"column_{index + 1}"
        for index, value in enumerate(non_empty[0])
    ]
    width = len(headers)
    data_rows = []
    for row in non_empty[1:]:
        normalized = [*row[:width], *([""] * max(0, width - len(row)))]
        if any(cell for cell in normalized):
            data_rows.append(normalized[:width])
    return headers, data_rows


def _spreadsheet_cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _spreadsheet_rows_to_text(
    *,
    headers: list[str],
    rows: list[list[str]],
    file_name: str,
    max_rows: int = 200,
) -> str:
    if not headers or not rows:
        return ""
    title = file_name or "spreadsheet"
    lines = [
        f"Spreadsheet source: {title}",
        "Columns: " + " | ".join(headers),
    ]
    for row_index, row in enumerate(rows[:max_rows], start=1):
        pairs = [
            f"{header}={row[column_index]}"
            for column_index, header in enumerate(headers)
            if column_index < len(row) and row[column_index]
        ]
        if pairs:
            lines.append(f"Row {row_index}: " + "; ".join(pairs))
    if len(rows) > max_rows:
        lines.append(f"Rows omitted after {max_rows}: {len(rows) - max_rows}")
    return _normalize_text("\n".join(lines))


def _contextual_source_text(
    *,
    request: OnboardingSourceIngestionRequest,
    index: int,
    chunk: str,
) -> str:
    payload = dict(request.source_payload)
    hints = [
        _optional_str(payload.get(key))
        for key in ("url", "file_name", "channel", "channel_id", "title")
    ]
    lines = [
        "Contextual source unit",
        f"Source kind: {request.source_kind}",
        f"Source ref: {request.source_ref}",
        f"Chunk index: {index}",
    ]
    for hint in hints:
        if hint:
            lines.append(f"Source hint: {hint}")
    lines.extend(["Evidence text:", chunk])
    return _normalize_text("\n".join(lines))


def _source_text_with_llm_context(
    *,
    context: str,
    source_text: str,
) -> str:
    return _normalize_text(
        "\n".join(
            [
                "LLM contextualized source unit",
                "LLM retrieval context:",
                context,
                "Original contextual source unit:",
                source_text,
            ]
        )
    )


def _preview(text: str) -> str:
    return text[:500]


def _unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _media_kind(value: Any) -> MediaKind | None:
    normalized = str(value or "").strip().lower()
    if normalized in {"photo", "image", "img"}:
        return "image"
    if normalized in {"video", "video_note"}:
        return "video"
    if normalized in {"audio", "voice"}:
        return "audio"
    if normalized in {"document", "file", "pdf"}:
        return "document"
    return None


def _content_type_for_name(value: str) -> str | None:
    content_type, _encoding = mimetypes.guess_type(value)
    return content_type


def _slug_for_ref(value: str) -> str:
    normalized = "".join(
        character.lower() if character.isalnum() else "-"
        for character in value.strip()
    )
    return "-".join(part for part in normalized.split("-") if part) or "item"


def _optional_str(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value)
